"""
HSD-ES API Data Fetcher with Excel Export
Fetches article data and status transitions, exports to Excel with merged cells
"""

import requests
from requests_kerberos import HTTPKerberosAuth, OPTIONAL
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import json
import os
import re
from collections import defaultdict

BASE_URL = 'https://hsdes-api.intel.com/rest/article'


def calculate_duration_from_date(date_str):
    """Calculate duration from given date to current date"""
    if not date_str:
        return ''
    
    try:
        # Parse the date string (format: 2025-09-09 11:43:10.747)
        if '.' in date_str:
            date_part = date_str.split('.')[0]  # Remove microseconds
        else:
            date_part = date_str
            
        transition_date = datetime.strptime(date_part, '%Y-%m-%d %H:%M:%S')
        current_date = datetime.now()
        
        # Calculate the difference
        time_diff = current_date - transition_date
        
        days = time_diff.days
        hours = time_diff.seconds // 3600
        
        if days > 0:
            return f"{days} Days {hours} Hours"
        else:
            return f"{hours} Hours"
            
    except Exception as e:
        print(f"Error calculating duration for date '{date_str}': {e}")
        return ''


def parse_time_spent_to_hours(time_str):
    """Parse time spent string to hours (float)"""
    if not time_str or time_str.strip() == '':
        return 0.0
    
    # Handle "< 1 Hour" case - treat as 1 hour
    if '<' in time_str and 'hour' in time_str.lower():
        return 1.0
    
    total_hours = 0.0
    time_str = str(time_str).strip().lower()
    
    # Extract days and hours
    days_match = re.search(r'(\d+)\s*days?', time_str)
    hours_match = re.search(r'(\d+)\s*hours?', time_str)
    
    if days_match:
        total_hours += int(days_match.group(1)) * 24
    
    if hours_match:
        total_hours += int(hours_match.group(1))
    
    return total_hours


def categorize_time_by_priority(hours, priority):
    """Categorize time into appropriate bucket based on priority"""
    days = hours / 24
    
    if priority == 'P1':
        if days < 2:
            return '< 2 days'
        elif days <= 5:
            return '2 to 5 days'
        else:
            return '> 5 days'
    
    elif priority == 'P2':
        if days < 5:
            return '< 5 days'
        elif days <= 7:
            return '6 to 7 days'
        else:
            return '> 7 days'
    
    elif priority == 'P3':
        if days < 10:
            return '< 10 days'
        elif days <= 15:
            return '10 to 15 days'
        else:
            return '> 15 days'
    
    elif priority == 'P4':
        if days < 15:
            return '< 15 days'
        else:
            return '> 15 days'
    
    return 'Unknown'


def get_rejected_articles_summary(article_data_list):
    """Get summary of rejected articles with article IDs
    
    Args:
        article_data_list: List of article data from API
        
    Returns:
        list: List of rejected article data with ID and rejection reason
    """
    rejected_articles = []
    
    for article_data in article_data_list:
        if not article_data:
            continue
            
        article_id = article_data.get('id', 'Unknown')
        is_rejected = False
        rejection_reason = ""
        
        # Check current status
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
            rejection_reason = article_data.get('status', '')
        
        # Check transitions for rejected status
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                if not rejection_reason:  # Use first rejected status found
                    rejection_reason = trans.get('status', '')
                break
        
        if is_rejected:
            rejected_articles.append({
                'id': article_id,
                'reason': rejection_reason,
                'priority': article_data.get('priority', '')
            })
    
    return rejected_articles


def analyze_open_new_to_ack_triage_transitions_from_api_data(article_data_list):
    """Analyze transitions from API data using the working method from reference script
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets
    """
    print(f"Analyzing {len(article_data_list)} articles for transition data...")
    
    # Initialize time buckets (from working script)
    time_buckets = {
        'P1': {'< 2 days': 0, '2 to 5 days': 0, '> 5 days': 0},
        'P2': {'< 5 days': 0, '6 to 7 days': 0, '> 7 days': 0},
        'P3': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P4': {'< 15 days': 0, '> 15 days': 0}
    }
    
    # Process each article (from working script logic)
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles - check ALL transitions for any rejected status
        is_rejected = False
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
        
        # Also check all transitions for rejected status
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status in transitions")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        article_id = article_data.get('id', 'Unknown')
        
        # Extract priority level (P1, P2, P3, P4) from format like "P1-SHOWSTOPPER"
        priority = None
        if priority_raw.startswith('P'):
            # Extract just the P1, P2, P3, P4 part
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]  # Get "P1" from "P1-SHOWSTOPPER"
            else:
                # Handle cases like "P1", "P2", etc.
                if len(priority_raw) >= 2 and priority_raw[1].isdigit():
                    priority = priority_raw[:2]  # Get "P1" from "P1xxx"
                else:
                    priority = priority_raw
        
        # Debug: Print priority extraction
        if priority not in ['P1', 'P2', 'P3', 'P4']:
            print(f"  DEBUG: Article {article_id} - Raw priority: '{priority_raw}' -> Extracted: '{priority}' (SKIPPED)")
            continue
        
        print(f"  DEBUG: Article {article_id} - Priority: {priority}")
            
        transitions = article_data['transitions']
        if not transitions:
            continue
        
        # Sort transitions by date (oldest first) - from working script
        sorted_transitions = []
        for trans in transitions:
            date = parse_date(trans.get('updated_date', ''))
            if date:
                sorted_transitions.append((date, trans))
        
        sorted_transitions.sort(key=lambda x: x[0])
        
        # Find first start status (open.new)
        start_date = None
        for date, trans in sorted_transitions:
            status = trans.get('status', '').lower()
            if status == 'open.new':
                start_date = date
                break
        
        if not start_date:
            continue
        
        # Find first end status after start (acknowledged, triage, or awaiting_submitter, not debug)
        end_date = None
        for date, trans in sorted_transitions:
            if date > start_date:
                status = trans.get('status', '').lower()
                if status == 'open.acknowledged' or status == 'open.triage' or status == 'open.awaiting_submitter':
                    end_date = date
                    break
        
        if not end_date:
            continue
        
        # Calculate days
        days = (end_date - start_date).days
        
        # Categorize into time buckets (fixed logic from working script)
        if priority == 'P1':
            if days < 2:
                time_buckets['P1']['< 2 days'] += 1
            elif days >= 2 and days <= 5:
                time_buckets['P1']['2 to 5 days'] += 1
            else:
                time_buckets['P1']['> 5 days'] += 1
        elif priority == 'P2':
            if days < 5:
                time_buckets['P2']['< 5 days'] += 1
            elif days >= 5 and days <= 7:  # Fixed: was >= 6, now >= 5 to remove gap
                time_buckets['P2']['6 to 7 days'] += 1
            else:
                time_buckets['P2']['> 7 days'] += 1
        elif priority == 'P3':
            if days < 10:
                time_buckets['P3']['< 10 days'] += 1
            elif days >= 10 and days <= 15:
                time_buckets['P3']['10 to 15 days'] += 1
            else:
                time_buckets['P3']['> 15 days'] += 1
        elif priority == 'P4':
            if days < 15:
                time_buckets['P4']['< 15 days'] += 1
            else:
                time_buckets['P4']['> 15 days'] += 1
    
    # Print final results summary (from working script)
    print(f"Analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Open.new to Open.acknowledged/Open.triage/Open.awaiting_submitter: {total_bugs} bugs analyzed")
    
    return time_buckets


def parse_time_spent_to_hours(time_str):
    """Parse time spent string to total hours
    
    Examples:
        '5 Days 23 Hours' -> 143 hours
        '< 1 Hour' -> 0.5 hours
        '28 Days 3 Hours' -> 675 hours
    """
    if not time_str:
        return 0
    
    time_str = time_str.strip().lower()
    
    # Handle '< 1 hour' cases
    if '< 1 hour' in time_str:
        return 0.5
    
    total_hours = 0
    
    # Extract days
    if 'day' in time_str:
        import re
        days_match = re.search(r'(\d+)\s+days?', time_str)
        if days_match:
            days = int(days_match.group(1))
            total_hours += days * 24
    
    # Extract hours
    if 'hour' in time_str:
        import re
        hours_match = re.search(r'(\d+)\s+hours?', time_str)
        if hours_match:
            hours = int(hours_match.group(1))
            total_hours += hours
    
    return total_hours


def analyze_start_to_end_transitions_from_api_data(article_data_list):
    """Analyze transitions from start statuses to end statuses
    
    Logic: 
    1. Find the LAST start status (open.new/acknowledged/triage) before any end status
    2. Find the LATEST end status (open.debug/promoted/root_caused/implemented) in the chain
    3. SUM all durations from the last start status to the latest end status (not including latest end)
    
    Example: open.new (13d19h) -> open.debug (9d3h) -> implemented -> verified -> complete
    - Last start: open.new (index 0)
    - Latest end: implemented (index 2)
    - Duration = open.new(13d19h) + open.debug(9d3h) = 22.9 days
    
    Start statuses: open.new, open.acknowledged, open.triage
    End statuses: open.debug, open.promoted, open.root_caused, implemented (EXACT match only)
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets
    """
    print(f"\\nAnalyzing {len(article_data_list)} articles for start-to-end transition data...")
    
    # Initialize time buckets for second analysis
    time_buckets = {
        'P1': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P2': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P3': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P4': {'< 15 days': 0, '> 15 days': 0}
    }
    
    # Define status groups - EXACT match for end statuses
    start_statuses = ['open.new', 'open.acknowledged', 'open.triage']
    end_statuses = ['open.debug', 'open.promoted', 'open.root_caused', 'implemented']
    
    # Process each article
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles
        is_rejected = False
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        
        # Extract priority level (P1, P2, P3, P4)
        priority = None
        if priority_raw.startswith('P'):
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]
            else:
                priority = priority_raw[:2] if len(priority_raw) >= 2 else priority_raw
        
        if not priority or priority not in ['P1', 'P2', 'P3', 'P4']:
            continue
        
        if not transitions:
            continue
        
        # Find last start status and LATEST end status, then sum durations between them
        last_start_status = None
        last_start_index = -1
        latest_end_status = None
        latest_end_index = -1
        
        # First pass: find all start and end statuses
        for i, trans in enumerate(transitions):
            status = trans.get('status', '').lower()
            
            if status in start_statuses:
                # Keep updating - we want the LAST start before any end
                if latest_end_index == -1:  # Only if we haven't hit an end status yet
                    last_start_status = status
                    last_start_index = i
            elif status in end_statuses:
                # Keep updating - we want the LATEST end status (EXACT match only)
                latest_end_status = status
                latest_end_index = i
        
        # Skip if no valid start or end found
        if last_start_index == -1 or latest_end_index == -1:
            continue
        
        # Statuses that cause early stopping - stop at last end status BEFORE these
        # Because their durations are counted in other analyses
        early_stop_statuses = ['open.awaiting_submitter', 'implemented.await_user_verify']
        
        # Check if any early-stop status is in the path
        early_stop_index = -1
        early_stop_status_found = None
        for i in range(last_start_index, latest_end_index + 1):
            status = transitions[i].get('status', '').lower()
            if status in early_stop_statuses:
                early_stop_index = i
                early_stop_status_found = status
                break
        
        if early_stop_index != -1:
            # Find the last end status BEFORE the early-stop status
            new_end_index = -1
            new_end_status = None
            for i in range(last_start_index, early_stop_index):
                status = transitions[i].get('status', '').lower()
                if status in end_statuses:
                    new_end_index = i
                    new_end_status = status
            
            if new_end_index == -1:
                # No end status before early-stop status, skip this article
                print(f"  SKIPPED: Article {article_id} - No end status before {early_stop_status_found}")
                continue
            
            latest_end_index = new_end_index
            latest_end_status = new_end_status
            print(f"  NOTE: Article {article_id} - Stopped at {latest_end_status} (before {early_stop_status_found})")
        
        # Statuses to exclude from duration sum (but not skip article)
        excluded_statuses = ['open.promoted', 'open.awaiting_3rd_party']
        
        # Sum durations from last_start_index to latest_end_index (not including latest_end)
        # EXCLUDE open.promoted and open.awaiting_3rd_party durations
        total_hours = 0.0
        duration_parts = []
        excluded_parts = []
        for i in range(last_start_index, latest_end_index):
            status = transitions[i].get('status', '').lower()
            dur = transitions[i].get('duration', '')
            if dur:
                if status in excluded_statuses:
                    # Skip these statuses from the sum
                    excluded_parts.append(f"{status}({dur})")
                else:
                    duration_parts.append(f"{status}({dur})")
                    total_hours += parse_time_spent_to_hours(dur)
        
        # Convert hours to days
        days = total_hours / 24
        
        duration_str = ' + '.join(duration_parts) if duration_parts else 'N/A'
        excluded_str = ', '.join(excluded_parts) if excluded_parts else ''
        if excluded_str:
            print(f"  COUNTED: Article {article_id} ({priority}) - {days:.1f} days [{duration_str}] -> {latest_end_status} (excluded: {excluded_str})")
        else:
            print(f"  COUNTED: Article {article_id} ({priority}) - {days:.1f} days [{duration_str}] -> {latest_end_status}")
        
        # Categorize into time buckets
        if priority in ['P1', 'P2', 'P3']:
            if days < 10:
                time_buckets[priority]['< 10 days'] += 1
            elif days <= 15:
                time_buckets[priority]['10 to 15 days'] += 1
            else:
                time_buckets[priority]['> 15 days'] += 1
        elif priority == 'P4':
            if days < 15:
                time_buckets[priority]['< 15 days'] += 1
            else:
                time_buckets[priority]['> 15 days'] += 1
    
    # Print final results summary
    print(f"Start-to-end analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Start-to-end transitions: {total_bugs} bugs analyzed")
    
    return time_buckets


def analyze_awaiting_submitter_transitions_from_api_data(article_data_list):
    """Analyze transitions from open.awaiting_submitter to any next state
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets (same as first analysis)
    """
    print(f"\nAnalyzing {len(article_data_list)} articles for awaiting_submitter transition data...")
    
    # Initialize time buckets (same as first analysis)
    time_buckets = {
        'P1': {'< 2 days': 0, '2 to 5 days': 0, '> 5 days': 0},
        'P2': {'< 5 days': 0, '6 to 7 days': 0, '> 7 days': 0},
        'P3': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P4': {'< 15 days': 0, '> 15 days': 0}
    }
    
    # Process each article
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles
        is_rejected = False
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
        
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        
        # Extract priority level (P1, P2, P3, P4)
        priority = None
        if priority_raw.startswith('P'):
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]
            else:
                if len(priority_raw) >= 2 and priority_raw[1].isdigit():
                    priority = priority_raw[:2]
                else:
                    priority = priority_raw
        
        if priority not in ['P1', 'P2', 'P3', 'P4']:
            continue
            
        if not transitions:
            continue
        
        # Sort transitions by date (oldest first)
        sorted_transitions = []
        for trans in transitions:
            date = parse_date(trans.get('updated_date', ''))
            if date:
                sorted_transitions.append((date, trans))
        
        sorted_transitions.sort(key=lambda x: x[0])
        
        # Find first open.awaiting_submitter status
        start_date = None
        for date, trans in sorted_transitions:
            status = trans.get('status', '').lower()
            if status == 'open.awaiting_submitter':
                start_date = date
                break
        
        if not start_date:
            continue
        
        # Find first status after open.awaiting_submitter (any different status)
        end_date = None
        for date, trans in sorted_transitions:
            if date > start_date:
                status = trans.get('status', '').lower()
                if status != 'open.awaiting_submitter':
                    end_date = date
                    break
        
        if not end_date:
            continue
        
        # Calculate days
        days = (end_date - start_date).days
        
        print(f"  COUNTED: Article {article_id} ({priority}) - {days} days (awaiting_submitter to next state)")
        
        # Categorize into time buckets (same as first analysis)
        if priority == 'P1':
            if days < 2:
                time_buckets['P1']['< 2 days'] += 1
            elif days >= 2 and days <= 5:
                time_buckets['P1']['2 to 5 days'] += 1
            else:
                time_buckets['P1']['> 5 days'] += 1
        elif priority == 'P2':
            if days < 5:
                time_buckets['P2']['< 5 days'] += 1
            elif days >= 5 and days <= 7:
                time_buckets['P2']['6 to 7 days'] += 1
            else:
                time_buckets['P2']['> 7 days'] += 1
        elif priority == 'P3':
            if days < 10:
                time_buckets['P3']['< 10 days'] += 1
            elif days >= 10 and days <= 15:
                time_buckets['P3']['10 to 15 days'] += 1
            else:
                time_buckets['P3']['> 15 days'] += 1
        elif priority == 'P4':
            if days < 15:
                time_buckets['P4']['< 15 days'] += 1
            else:
                time_buckets['P4']['> 15 days'] += 1
    
    # Print final results summary
    print(f"Awaiting_submitter analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Open.awaiting_submitter to next state: {total_bugs} bugs analyzed")
    
    return time_buckets


def analyze_promoted_to_implemented_transitions_from_api_data(article_data_list):
    """Analyze transitions from open.promoted to any next state
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets (same as second analysis)
    """
    print(f"\nAnalyzing {len(article_data_list)} articles for promoted to next state transition data...")
    
    # Initialize time buckets (same as second analysis)
    time_buckets = {
        'P1': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P2': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P3': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P4': {'< 15 days': 0, '> 15 days': 0}
    }
    
    # Process each article
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles
        is_rejected = False
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
        
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        
        # Extract priority level (P1, P2, P3, P4)
        priority = None
        if priority_raw.startswith('P'):
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]
            else:
                if len(priority_raw) >= 2 and priority_raw[1].isdigit():
                    priority = priority_raw[:2]
                else:
                    priority = priority_raw
        
        if priority not in ['P1', 'P2', 'P3', 'P4']:
            continue
            
        if not transitions:
            continue
        
        # Sort transitions by date (oldest first)
        sorted_transitions = []
        for trans in transitions:
            date = parse_date(trans.get('updated_date', ''))
            if date:
                sorted_transitions.append((date, trans))
        
        sorted_transitions.sort(key=lambda x: x[0])
        
        # Find first open.promoted status
        start_date = None
        for date, trans in sorted_transitions:
            status = trans.get('status', '').lower()
            if status == 'open.promoted':
                start_date = date
                break
        
        if not start_date:
            continue
        
        # Find first status after open.promoted (any different status)
        end_date = None
        end_status_found = None
        for date, trans in sorted_transitions:
            if date > start_date:
                status = trans.get('status', '').lower()
                if status != 'open.promoted':
                    end_date = date
                    end_status_found = status
                    break
        
        if not end_date:
            continue
        
        # Calculate days
        days = (end_date - start_date).days
        
        print(f"  COUNTED: Article {article_id} ({priority}) - {days} days (promoted to {end_status_found})")
        
        # Categorize into time buckets (same as second analysis)
        if priority in ['P1', 'P2', 'P3']:
            if days < 10:
                time_buckets[priority]['< 10 days'] += 1
            elif days <= 15:
                time_buckets[priority]['10 to 15 days'] += 1
            else:
                time_buckets[priority]['> 15 days'] += 1
        elif priority == 'P4':
            if days < 15:
                time_buckets[priority]['< 15 days'] += 1
            else:
                time_buckets[priority]['> 15 days'] += 1
    
    # Print final results summary
    print(f"Promoted to next state analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Open.promoted to next state: {total_bugs} bugs analyzed")
    
    return time_buckets


def analyze_await_user_verify_transitions_from_api_data(article_data_list):
    """Analyze transitions from implemented/implemented.await_user_verify to implemented/verified
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets (< 7, 7-10, > 10 days)
    """
    print(f"\nAnalyzing {len(article_data_list)} articles for implemented/await_user_verify to implemented/verified transition data...")
    
    # Initialize time buckets (< 7, 7-10, > 10 days for P1/P2/P3, < 10, > 10 for P4)
    time_buckets = {
        'P1': {'< 7 days': 0, '7 to 10 days': 0, '> 10 days': 0},
        'P2': {'< 7 days': 0, '7 to 10 days': 0, '> 10 days': 0},
        'P3': {'< 7 days': 0, '7 to 10 days': 0, '> 10 days': 0},
        'P4': {'< 10 days': 0, '> 10 days': 0}
    }
    
    # Define valid start and end statuses
    valid_start_statuses = ['implemented', 'implemented.await_user_verify']
    valid_end_statuses = ['implemented', 'verified']
    
    # Process each article
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles
        is_rejected = False
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
        
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        
        # Extract priority level (P1, P2, P3, P4)
        priority = None
        if priority_raw.startswith('P'):
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]
            else:
                if len(priority_raw) >= 2 and priority_raw[1].isdigit():
                    priority = priority_raw[:2]
                else:
                    priority = priority_raw
        
        if priority not in ['P1', 'P2', 'P3', 'P4']:
            continue
            
        if not transitions:
            continue
        
        # Sort transitions by date (oldest first)
        sorted_transitions = []
        for trans in transitions:
            date = parse_date(trans.get('updated_date', ''))
            if date:
                sorted_transitions.append((date, trans))
        
        sorted_transitions.sort(key=lambda x: x[0])
        
        # Find first implemented or implemented.await_user_verify status
        start_date = None
        start_status_found = None
        for date, trans in sorted_transitions:
            status = trans.get('status', '').lower()
            if status in valid_start_statuses:
                start_date = date
                start_status_found = status
                break
        
        if not start_date:
            continue
        
        # Find first implemented or verified status after start (must be different from start status)
        end_date = None
        end_status_found = None
        for date, trans in sorted_transitions:
            if date > start_date:
                status = trans.get('status', '').lower()
                if status in valid_end_statuses and status != start_status_found:
                    end_date = date
                    end_status_found = status
                    break
        
        if not end_date:
            continue
        
        # Calculate days (using total_seconds for fractional days)
        time_diff = end_date - start_date
        days = time_diff.total_seconds() / (24 * 3600)
        
        print(f"  COUNTED: Article {article_id} ({priority}) - {days:.2f} days ({start_status_found} to {end_status_found})")
        
        # Categorize into time buckets (< 7, 7-10, > 10 days for P1/P2/P3, < 10, > 10 for P4)
        if priority in ['P1', 'P2', 'P3']:
            if days < 7:
                time_buckets[priority]['< 7 days'] += 1
            elif days <= 10:
                time_buckets[priority]['7 to 10 days'] += 1
            else:
                time_buckets[priority]['> 10 days'] += 1
        elif priority == 'P4':
            if days < 10:
                time_buckets['P4']['< 10 days'] += 1
            else:
                time_buckets['P4']['> 10 days'] += 1
    
    # Print final results summary
    print(f"Implemented/await_user_verify to implemented/verified analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Implemented/await_user_verify to implemented/verified: {total_bugs} bugs analyzed")
    
    return time_buckets


def analyze_any_to_complete_product_changed_transitions_from_api_data(article_data_list):
    """Analyze transitions from any state to complete.product_changed
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets (same as second analysis)
    """
    print(f"\nAnalyzing {len(article_data_list)} articles for any state to complete.product_changed transition data...")
    
    # Initialize time buckets (same as second analysis - 10/15 day buckets)
    time_buckets = {
        'P1': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P2': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P3': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P4': {'< 15 days': 0, '> 15 days': 0}
    }
    
    # Process each article
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles
        is_rejected = False
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
        
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        
        # Extract priority level (P1, P2, P3, P4)
        priority = None
        if priority_raw.startswith('P'):
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]
            else:
                if len(priority_raw) >= 2 and priority_raw[1].isdigit():
                    priority = priority_raw[:2]
                else:
                    priority = priority_raw
        
        if priority not in ['P1', 'P2', 'P3', 'P4']:
            continue
            
        if not transitions:
            continue
        
        # Sort transitions by date (oldest first)
        sorted_transitions = []
        for trans in transitions:
            date = parse_date(trans.get('updated_date', ''))
            if date:
                sorted_transitions.append((date, trans))
        
        sorted_transitions.sort(key=lambda x: x[0])
        
        # Find first complete.product_changed status
        complete_date = None
        complete_index = -1
        for i, (date, trans) in enumerate(sorted_transitions):
            status = trans.get('status', '').lower()
            if status == 'complete.product_changed':
                complete_date = date
                complete_index = i
                break
        
        if not complete_date:
            continue
        
        # Find the state immediately before complete.product_changed
        if complete_index == 0:
            # complete.product_changed is the first status, skip
            continue
        
        # Get the state just before complete.product_changed
        prev_date, prev_trans = sorted_transitions[complete_index - 1]
        prev_status = prev_trans.get('status', '').lower()
        
        # Calculate days from previous state to complete.product_changed (using total_seconds for fractional days)
        time_diff = complete_date - prev_date
        days = time_diff.total_seconds() / (24 * 3600)
        
        print(f"  COUNTED: Article {article_id} ({priority}) - {days:.2f} days ({prev_status} to complete.product_changed)")
        
        # Categorize into time buckets (same as second analysis - 10/15 day buckets)
        if priority in ['P1', 'P2', 'P3']:
            if days < 10:
                time_buckets[priority]['< 10 days'] += 1
            elif days <= 15:
                time_buckets[priority]['10 to 15 days'] += 1
            else:
                time_buckets[priority]['> 15 days'] += 1
        elif priority == 'P4':
            if days < 15:
                time_buckets['P4']['< 15 days'] += 1
            else:
                time_buckets['P4']['> 15 days'] += 1
    
    # Print final results summary
    print(f"Any state to complete.product_changed analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Any state to complete.product_changed: {total_bugs} bugs analyzed")
    
    return time_buckets


def analyze_new_to_await_user_verify_transitions_from_api_data(article_data_list):
    """Analyze transitions from open.new to implemented.await_user_verify
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets (same as first analysis)
    """
    print(f"\nAnalyzing {len(article_data_list)} articles for open.new to await_user_verify transition data...")
    
    # Initialize time buckets (same as first analysis)
    time_buckets = {
        'P1': {'< 2 days': 0, '2 to 5 days': 0, '> 5 days': 0},
        'P2': {'< 5 days': 0, '6 to 7 days': 0, '> 7 days': 0},
        'P3': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P4': {'< 15 days': 0, '> 15 days': 0}
    }
    
    # Process each article
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles
        is_rejected = False
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
        
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        
        # Extract priority level (P1, P2, P3, P4)
        priority = None
        if priority_raw.startswith('P'):
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]
            else:
                if len(priority_raw) >= 2 and priority_raw[1].isdigit():
                    priority = priority_raw[:2]
                else:
                    priority = priority_raw
        
        if priority not in ['P1', 'P2', 'P3', 'P4']:
            continue
            
        if not transitions:
            continue
        
        # Sort transitions by date (oldest first)
        sorted_transitions = []
        for trans in transitions:
            date = parse_date(trans.get('updated_date', ''))
            if date:
                sorted_transitions.append((date, trans))
        
        sorted_transitions.sort(key=lambda x: x[0])
        
        # Find first open.new status
        start_date = None
        for date, trans in sorted_transitions:
            status = trans.get('status', '').lower()
            if status == 'open.new':
                start_date = date
                break
        
        if not start_date:
            continue
        
        # Find first implemented.await_user_verify after open.new
        end_date = None
        for date, trans in sorted_transitions:
            if date > start_date:
                status = trans.get('status', '').lower()
                if status == 'implemented.await_user_verify':
                    end_date = date
                    break
        
        if not end_date:
            continue
        
        # Calculate days (using total_seconds for fractional days)
        time_diff = end_date - start_date
        days = time_diff.total_seconds() / (24 * 3600)
        
        print(f"  COUNTED: Article {article_id} ({priority}) - {days:.2f} days (new to await_user_verify)")
        
        # Categorize into time buckets (same as first analysis)
        if priority == 'P1':
            if days < 2:
                time_buckets['P1']['< 2 days'] += 1
            elif days <= 5:
                time_buckets['P1']['2 to 5 days'] += 1
            else:
                time_buckets['P1']['> 5 days'] += 1
        elif priority == 'P2':
            if days < 5:
                time_buckets['P2']['< 5 days'] += 1
            elif days <= 7:
                time_buckets['P2']['6 to 7 days'] += 1
            else:
                time_buckets['P2']['> 7 days'] += 1
        elif priority == 'P3':
            if days < 10:
                time_buckets['P3']['< 10 days'] += 1
            elif days <= 15:
                time_buckets['P3']['10 to 15 days'] += 1
            else:
                time_buckets['P3']['> 15 days'] += 1
        elif priority == 'P4':
            if days < 15:
                time_buckets['P4']['< 15 days'] += 1
            else:
                time_buckets['P4']['> 15 days'] += 1
    
    # Print final results summary
    print(f"Open.new to await_user_verify analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Open.new to implemented.await_user_verify: {total_bugs} bugs analyzed")
    
    return time_buckets


def analyze_any_to_await_user_verify_transitions_from_api_data(article_data_list):
    """Analyze transitions from any state (NOT open.new) to implemented.await_user_verify
    
    Args:
        article_data_list: List of article data with transitions from API
        
    Returns:
        dict: Analysis results grouped by priority with time buckets (same as second analysis)
    """
    print(f"\nAnalyzing {len(article_data_list)} articles for any state (not new) to await_user_verify transition data...")
    
    # Initialize time buckets (same as second analysis - 10/15 day buckets)
    time_buckets = {
        'P1': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P2': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P3': {'< 10 days': 0, '10 to 15 days': 0, '> 15 days': 0},
        'P4': {'< 15 days': 0, '> 15 days': 0}
    }
    
    # Process each article
    for article_data in article_data_list:
        if not article_data or 'transitions' not in article_data:
            continue
        
        article_id = article_data.get('id', 'Unknown')
        
        # Skip rejected articles
        is_rejected = False
        current_status = article_data.get('status', '').lower()
        if 'rejected' in current_status:
            is_rejected = True
        
        transitions = article_data.get('transitions', [])
        for trans in transitions:
            trans_status = trans.get('status', '').lower()
            if 'rejected' in trans_status:
                is_rejected = True
                break
        
        if is_rejected:
            print(f"  SKIPPED: Article {article_id} - Has rejected status")
            continue
            
        priority_raw = article_data.get('priority', '').upper()
        
        # Extract priority level (P1, P2, P3, P4)
        priority = None
        if priority_raw.startswith('P'):
            if '-' in priority_raw:
                priority = priority_raw.split('-')[0]
            else:
                if len(priority_raw) >= 2 and priority_raw[1].isdigit():
                    priority = priority_raw[:2]
                else:
                    priority = priority_raw
        
        if priority not in ['P1', 'P2', 'P3', 'P4']:
            continue
            
        if not transitions:
            continue
        
        # Sort transitions by date (oldest first)
        sorted_transitions = []
        for trans in transitions:
            date = parse_date(trans.get('updated_date', ''))
            if date:
                sorted_transitions.append((date, trans))
        
        sorted_transitions.sort(key=lambda x: x[0])
        
        # Find first implemented.await_user_verify status
        await_verify_date = None
        await_verify_index = -1
        for i, (date, trans) in enumerate(sorted_transitions):
            status = trans.get('status', '').lower()
            if status == 'implemented.await_user_verify':
                await_verify_date = date
                await_verify_index = i
                break
        
        if not await_verify_date:
            continue
        
        # Find the state immediately before await_user_verify (must NOT be open.new)
        if await_verify_index == 0:
            # await_user_verify is the first status, skip
            continue
        
        # Get the state just before await_user_verify
        prev_date, prev_trans = sorted_transitions[await_verify_index - 1]
        prev_status = prev_trans.get('status', '').lower()
        
        # Skip if previous state is open.new (those are counted in fifth analysis)
        if prev_status == 'open.new':
            print(f"  SKIPPED: Article {article_id} - Previous state is open.new (counted in fifth analysis)")
            continue
        
        # Calculate days from previous state to await_user_verify (using total_seconds for fractional days)
        time_diff = await_verify_date - prev_date
        days = time_diff.total_seconds() / (24 * 3600)
        
        print(f"  COUNTED: Article {article_id} ({priority}) - {days:.2f} days ({prev_status} to await_user_verify)")
        
        # Categorize into time buckets (same as second analysis - 10/15 day buckets)
        if priority in ['P1', 'P2', 'P3']:
            if days < 10:
                time_buckets[priority]['< 10 days'] += 1
            elif days <= 15:
                time_buckets[priority]['10 to 15 days'] += 1
            else:
                time_buckets[priority]['> 15 days'] += 1
        elif priority == 'P4':
            if days < 15:
                time_buckets['P4']['< 15 days'] += 1
            else:
                time_buckets['P4']['> 15 days'] += 1
    
    # Print final results summary
    print(f"Any state (not new) to await_user_verify analysis complete!")
    total_bugs = 0
    for priority, buckets in time_buckets.items():
        priority_total = sum(buckets.values())
        total_bugs += priority_total
        if priority_total > 0:
            print(f"  {priority}: {buckets} (total: {priority_total})")
    print(f"  Any state (not open.new) to implemented.await_user_verify: {total_bugs} bugs analyzed")
    
    return time_buckets


def parse_date(date_str):
    """Parse date string to datetime object (from working script)"""
    if not date_str:
        return None
    
    try:
        # Handle different date formats
        if '.' in date_str:
            date_part = date_str.split('.')[0]  # Remove microseconds
        else:
            date_part = date_str
            
        return datetime.strptime(date_part, '%Y-%m-%d %H:%M:%S')
    except:
        return None


def parse_excel_date(date_str):
    """Parse date string from Excel to datetime object"""
    if not date_str:
        return None
    
    try:
        # Handle different date formats
        if '.' in date_str:
            date_part = date_str.split('.')[0]  # Remove microseconds
        else:
            date_part = date_str
            
        return datetime.strptime(date_part, '%Y-%m-%d %H:%M:%S')
    except:
        return None


def create_wcl_bugs_transition_graph_sheet(excel_file_path, article_data_list, platform_name="WCL"):
    """Create Bugs_transition_graph sheet with rejected articles summary and transition analysis
    
    Args:
        excel_file_path (str): Path to the existing Excel file to add the sheet to
        article_data_list (list): Article data collected from API
        platform_name (str): Platform name for sheet and table titles (e.g., 'WCL', 'PTL-H')
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        
        # Get rejected articles summary first
        rejected_articles = get_rejected_articles_summary(article_data_list)
        
        # Run all three analyses
        print("Running first analysis: open.new to acknowledged/triage...")
        first_analysis = analyze_open_new_to_ack_triage_transitions_from_api_data(article_data_list)
        
        print("\\nRunning second analysis: start statuses to implemented...")
        second_analysis = analyze_start_to_end_transitions_from_api_data(article_data_list)
        
        print("\\nRunning third analysis: awaiting_submitter to next state...")
        third_analysis = analyze_awaiting_submitter_transitions_from_api_data(article_data_list)
        
        print("\\nRunning fourth analysis: promoted to implemented/awaiting_3rd_party...")
        fourth_analysis = analyze_promoted_to_implemented_transitions_from_api_data(article_data_list)
        
        print("\\nRunning fifth analysis: open.new to await_user_verify...")
        fifth_analysis = analyze_new_to_await_user_verify_transitions_from_api_data(article_data_list)
        
        print("\\nRunning sixth analysis: any state (not new) to await_user_verify...")
        sixth_analysis = analyze_any_to_await_user_verify_transitions_from_api_data(article_data_list)
        
        print("\\nRunning seventh analysis: await_user_verify to implemented/verified/complete.product_changed...")
        seventh_analysis = analyze_await_user_verify_transitions_from_api_data(article_data_list)
        
        print("\\nRunning eighth analysis: any state to complete.product_changed...")
        eighth_analysis = analyze_any_to_complete_product_changed_transitions_from_api_data(article_data_list)
        
        if not first_analysis and not second_analysis and not third_analysis and not fourth_analysis and not fifth_analysis and not sixth_analysis and not seventh_analysis and not eighth_analysis:
            print("ERROR: Failed to analyze transitions")
            return False
        
        # Load existing workbook
        wb = load_workbook(excel_file_path)
        
        # Create sheet name based on platform
        sheet_name = f'{platform_name}_Bugs_transition_graph'
        
        # Create new sheet or clear existing one
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        ws = wb.create_sheet(sheet_name)
        
        # Define styling
        header_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        rejected_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        current_row = 1
        
        # CREATE REJECTED ARTICLES SUMMARY TABLE AT TOP
        if rejected_articles:
            # Title
            title_cell = ws.cell(row=current_row, column=2, value=f"REJECTED ARTICLES SUMMARY (Total: {len(rejected_articles)})")
            title_cell.font = title_font
            title_cell.fill = rejected_fill
            title_cell.alignment = center_alignment
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=4)
            current_row += 1
            
            # Headers
            headers = ['Article ID', 'Priority', 'Rejection Reason']
            for i, header in enumerate(headers):
                cell = ws.cell(row=current_row, column=2+i, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = center_alignment
            current_row += 1
            
            # Data rows with hyperlinks
            for article in rejected_articles:
                # Article ID with hyperlink
                hyperlink_formula = f'=HYPERLINK("https://hsdes.intel.com/resource/{article["id"]}", "{article["id"]}")'  
                id_cell = ws.cell(row=current_row, column=2, value=hyperlink_formula)
                id_cell.border = border
                
                ws.cell(row=current_row, column=3, value=article['priority']).border = border
                ws.cell(row=current_row, column=4, value=article['reason']).border = border
                current_row += 1
            
            # Add spacing
            current_row += 2
        
        # Now adjust table positions to start after rejected articles summary
        table_start_row = current_row
        
        # Define time buckets for each analysis
        first_buckets = {
            'P1': ['< 2 days', '2 to 5 days', '> 5 days'],
            'P2': ['< 5 days', '6 to 7 days', '> 7 days'],
            'P3': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P4': ['< 15 days', '> 15 days']
        }
        
        second_buckets = {
            'P1': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P2': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P3': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P4': ['< 15 days', '> 15 days']
        }
        
        # Third buckets same as first
        third_buckets = {
            'P1': ['< 2 days', '2 to 5 days', '> 5 days'],
            'P2': ['< 5 days', '6 to 7 days', '> 7 days'],
            'P3': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P4': ['< 15 days', '> 15 days']
        }
        
        # Fourth buckets same as second
        fourth_buckets = {
            'P1': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P2': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P3': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P4': ['< 15 days', '> 15 days']
        }
        
        # Fifth buckets same as first (open.new to await_user_verify)
        fifth_buckets = {
            'P1': ['< 2 days', '2 to 5 days', '> 5 days'],
            'P2': ['< 5 days', '6 to 7 days', '> 7 days'],
            'P3': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P4': ['< 15 days', '> 15 days']
        }
        
        # Sixth buckets same as second (any state not new to await_user_verify)
        sixth_buckets = {
            'P1': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P2': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P3': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P4': ['< 15 days', '> 15 days']
        }
        
        # Seventh buckets (< 7, 7-10, > 10 days for P1/P2/P3, < 10, > 10 for P4)
        seventh_buckets = {
            'P1': ['< 7 days', '7 to 10 days', '> 10 days'],
            'P2': ['< 7 days', '7 to 10 days', '> 10 days'],
            'P3': ['< 7 days', '7 to 10 days', '> 10 days'],
            'P4': ['< 10 days', '> 10 days']
        }
        
        # Eighth buckets same as second (any state to complete.product_changed)
        eighth_buckets = {
            'P1': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P2': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P3': ['< 10 days', '10 to 15 days', '> 15 days'],
            'P4': ['< 15 days', '> 15 days']
        }
        
        def create_table_set(ws, analysis_data, buckets_def, positions, title_suffix, 
                            header_font, title_font, header_fill, border, center_alignment, platform):
            """Helper function to create a set of 4 tables"""
            for priority in ['P1', 'P2', 'P3', 'P4']:
                start_row, start_col = positions[priority]
                
                # Table title
                title_cell = ws.cell(row=start_row, column=start_col, 
                                   value=f"{platform} {priority} bugs - {title_suffix}")
                title_cell.font = title_font
                title_cell.alignment = center_alignment
                
                # Merge title across 2 columns
                ws.merge_cells(start_row=start_row, start_column=start_col, 
                              end_row=start_row, end_column=start_col + 1)
                
                # Table headers
                header_row = start_row + 1
                time_header = ws.cell(row=header_row, column=start_col, value="Time")
                count_header = ws.cell(row=header_row, column=start_col + 1, value="Bug count")
                
                for header_cell in [time_header, count_header]:
                    header_cell.font = header_font
                    header_cell.fill = header_fill
                    header_cell.border = border
                    header_cell.alignment = center_alignment
                
                # Table data
                buckets = buckets_def[priority]
                results = analysis_data.get(priority, {})
                
                for i, time_bucket in enumerate(buckets):
                    data_row = header_row + 1 + i
                    count = results.get(time_bucket, 0)
                    
                    time_cell = ws.cell(row=data_row, column=start_col, value=time_bucket)
                    count_cell = ws.cell(row=data_row, column=start_col + 1, value=count)
                    
                    for cell in [time_cell, count_cell]:
                        cell.border = border
                        cell.alignment = center_alignment
        
        # FIRST SET OF TABLES
        first_table_positions = {
            'P1': (table_start_row, 2),   # B(start_row)
            'P2': (table_start_row, 6),   # F(start_row)
            'P3': (table_start_row, 10),  # J(start_row)
            'P4': (table_start_row, 14)   # N(start_row)
        }
        
        # Define fonts here for the helper function
        title_font = Font(bold=True, size=12)
        
        create_table_set(ws, first_analysis, first_buckets, first_table_positions,
                        "Open.new to Open.acknowledged/Open.triage/Open.awaiting_submitter", 
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # SECOND SET OF TABLES (with space for graphs)
        second_table_start_row = table_start_row + 15  # Leave space for graphs
        second_table_positions = {
            'P1': (second_table_start_row, 2),   # B(start_row + 15)
            'P2': (second_table_start_row, 6),   # F(start_row + 15)
            'P3': (second_table_start_row, 10),  # J(start_row + 15)
            'P4': (second_table_start_row, 14)   # N(start_row + 15)
        }
        
        create_table_set(ws, second_analysis, second_buckets, second_table_positions,
                        "Open.new/Open.acknowledged/Open.triage to Open.debug/Open.promoted/Open.root_caused/Implemented",
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # THIRD SET OF TABLES (with space for graphs)
        third_table_start_row = table_start_row + 30  # Leave space for second set graphs
        third_table_positions = {
            'P1': (third_table_start_row, 2),   # B(start_row + 30)
            'P2': (third_table_start_row, 6),   # F(start_row + 30)
            'P3': (third_table_start_row, 10),  # J(start_row + 30)
            'P4': (third_table_start_row, 14)   # N(start_row + 30)
        }
        
        create_table_set(ws, third_analysis, third_buckets, third_table_positions,
                        "Open.awaiting_submitter to Next State",
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # FOURTH SET OF TABLES (with space for graphs)
        fourth_table_start_row = table_start_row + 45  # Leave space for third set graphs
        fourth_table_positions = {
            'P1': (fourth_table_start_row, 2),   # B(start_row + 45)
            'P2': (fourth_table_start_row, 6),   # F(start_row + 45)
            'P3': (fourth_table_start_row, 10),  # J(start_row + 45)
            'P4': (fourth_table_start_row, 14)   # N(start_row + 45)
        }
        
        create_table_set(ws, fourth_analysis, fourth_buckets, fourth_table_positions,
                        "Open.promoted to Next State",
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # FIFTH SET OF TABLES (with space for graphs)
        fifth_table_start_row = table_start_row + 60  # Leave space for fourth set graphs
        fifth_table_positions = {
            'P1': (fifth_table_start_row, 2),   # B(start_row + 60)
            'P2': (fifth_table_start_row, 6),   # F(start_row + 60)
            'P3': (fifth_table_start_row, 10),  # J(start_row + 60)
            'P4': (fifth_table_start_row, 14)   # N(start_row + 60)
        }
        
        create_table_set(ws, fifth_analysis, fifth_buckets, fifth_table_positions,
                        "Open.new to Implemented.await_user_verify",
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # SIXTH SET OF TABLES (with space for graphs)
        sixth_table_start_row = table_start_row + 75  # Leave space for fifth set graphs
        sixth_table_positions = {
            'P1': (sixth_table_start_row, 2),   # B(start_row + 75)
            'P2': (sixth_table_start_row, 6),   # F(start_row + 75)
            'P3': (sixth_table_start_row, 10),  # J(start_row + 75)
            'P4': (sixth_table_start_row, 14)   # N(start_row + 75)
        }
        
        create_table_set(ws, sixth_analysis, sixth_buckets, sixth_table_positions,
                        "Any State (otherthan Open.new) to Implemented.await_user_verify",
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # SEVENTH SET OF TABLES (with space for graphs)
        seventh_table_start_row = table_start_row + 90  # Leave space for sixth set graphs
        seventh_table_positions = {
            'P1': (seventh_table_start_row, 2),   # B(start_row + 90)
            'P2': (seventh_table_start_row, 6),   # F(start_row + 90)
            'P3': (seventh_table_start_row, 10),  # J(start_row + 90)
            'P4': (seventh_table_start_row, 14)   # N(start_row + 90)
        }
        
        create_table_set(ws, seventh_analysis, seventh_buckets, seventh_table_positions,
                        "Implemented/Implemented.await_user_verify to Implemented/Verified",
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # EIGHTH SET OF TABLES (with space for graphs)
        eighth_table_start_row = table_start_row + 105  # Leave space for seventh set graphs
        eighth_table_positions = {
            'P1': (eighth_table_start_row, 2),   # B(start_row + 105)
            'P2': (eighth_table_start_row, 6),   # F(start_row + 105)
            'P3': (eighth_table_start_row, 10),  # J(start_row + 105)
            'P4': (eighth_table_start_row, 14)   # N(start_row + 105)
        }
        
        create_table_set(ws, eighth_analysis, eighth_buckets, eighth_table_positions,
                        "Any State to Complete.product_changed",
                        header_font, title_font, header_fill, border, center_alignment, platform_name)
        
        # Adjust column widths for better presentation
        for col in range(1, 18):  # A to R
            col_letter = get_column_letter(col)
            if col == 1:  # Column A - empty spacing
                ws.column_dimensions[col_letter].width = 2
            elif col in [2, 6, 10, 14]:  # Time columns (B, F, J, N)
                ws.column_dimensions[col_letter].width = 18
            elif col in [3, 7, 11, 15]:  # Bug count columns (C, G, K, O)
                ws.column_dimensions[col_letter].width = 12
            else:  # Spacing columns between tables
                ws.column_dimensions[col_letter].width = 3
        
        # Save the workbook
        wb.save(excel_file_path)
        wb.close()
        
        print("\\n" + "="*80)
        print("SUCCESS: WCL_Bugs_transition_graph sheet created successfully!")
        
        if rejected_articles:
            print(f"\\nREJECTED ARTICLES SUMMARY: {len(rejected_articles)} articles excluded")
            for article in rejected_articles[:5]:  # Show first 5
                print(f"   • {article['id']} ({article['priority']}) - {article['reason']}")
            if len(rejected_articles) > 5:
                print(f"   • ... and {len(rejected_articles) - 5} more (see sheet for full list)")
        
        print("\\nTransition Analysis Summary:")
        print("\\nFIRST ANALYSIS (Open.new to Open.acknowledged/Open.triage/Open.awaiting_submitter):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = first_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        
        print("\\nSECOND ANALYSIS (Open.new/ack/triage to Open.debug/promoted/root_caused/Implemented):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = second_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        
        print("\\nTHIRD ANALYSIS (Open.awaiting_submitter to Next State):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = third_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        
        print("\\nFOURTH ANALYSIS (Open.promoted to Next State):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = fourth_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        
        print("\\nFIFTH ANALYSIS (Open.new to Implemented.await_user_verify):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = fifth_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        
        print("\\nSIXTH ANALYSIS (Any State (otherthan Open.new) to Implemented.await_user_verify):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = sixth_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        
        print("\\nSEVENTH ANALYSIS (Implemented/Implemented.await_user_verify to Implemented/Verified):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = seventh_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        
        print("\\nEIGHTH ANALYSIS (Any State to Complete.product_changed):")
        for priority in ['P1', 'P2', 'P3', 'P4']:
            results = eighth_analysis.get(priority, {})
            total_bugs = sum(results.values())
            print(f"  {priority}: {total_bugs} bugs analyzed")
            for bucket, count in results.items():
                if count > 0:
                    print(f"    {bucket}: {count} bugs")
        print("="*80)
        
        return True
        
    except Exception as e:
        print(f"ERROR creating transition graph sheet: {e}")
        return False


def create_status_summary_sheet(excel_file_path, article_data_list, platform_name="WCL"):
    """Create a sheet with article status summary
    
    Args:
        excel_file_path (str): Path to the existing Excel file to add the sheet to
        article_data_list (list): Article data collected from API
        platform_name (str): Platform name for sheet title
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
        from collections import Counter
        
        # Load existing workbook
        wb = load_workbook(excel_file_path)
        
        # Create sheet name based on platform
        sheet_name = f'{platform_name}_Status_Summary'
        
        # Create new sheet or clear existing one
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        
        ws = wb.create_sheet(sheet_name)
        
        # Define styling
        header_font = Font(bold=True, size=11)
        title_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
        count_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        # ========== FIRST TABLE: Article ID with Current Status ==========
        current_row = 1
        
        # Title for first table
        title_cell = ws.cell(row=current_row, column=1, value=f"{platform_name} - Article Current Status")
        title_cell.font = title_font
        title_cell.alignment = center_alignment
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1
        
        # Headers for first table
        headers = ['Article ID', 'Priority', 'Current Status']
        for i, header in enumerate(headers):
            cell = ws.cell(row=current_row, column=1+i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1
        
        # Collect status data
        status_list = []
        for article_data in article_data_list:
            if not article_data:
                continue
            
            article_id = article_data.get('id', 'Unknown')
            priority = article_data.get('priority', 'Unknown')
            current_status = article_data.get('status', 'Unknown')
            
            status_list.append({
                'id': article_id,
                'priority': priority,
                'status': current_status
            })
            
            # Article ID with hyperlink
            hyperlink_formula = f'=HYPERLINK("https://hsdes.intel.com/resource/{article_id}", "{article_id}")'
            id_cell = ws.cell(row=current_row, column=1, value=hyperlink_formula)
            id_cell.border = border
            id_cell.alignment = center_alignment
            
            # Priority
            priority_cell = ws.cell(row=current_row, column=2, value=priority)
            priority_cell.border = border
            priority_cell.alignment = center_alignment
            
            # Current Status
            status_cell = ws.cell(row=current_row, column=3, value=current_status)
            status_cell.border = border
            status_cell.alignment = left_alignment
            
            current_row += 1
        
        # Add total count row
        total_cell = ws.cell(row=current_row, column=1, value="Total Articles:")
        total_cell.font = header_font
        total_cell.border = border
        ws.cell(row=current_row, column=2, value=len(status_list)).border = border
        ws.cell(row=current_row, column=3, value="").border = border
        current_row += 2
        
        # ========== SECOND TABLE: Status Count Summary ==========
        # Count articles by status
        status_counter = Counter([item['status'] for item in status_list])
        
        # Title for second table
        title_cell2 = ws.cell(row=current_row, column=1, value=f"{platform_name} - Status Distribution Summary")
        title_cell2.font = title_font
        title_cell2.alignment = center_alignment
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=3)
        current_row += 1
        
        # Headers for second table
        headers2 = ['Status', 'Article Count', 'Percentage']
        for i, header in enumerate(headers2):
            cell = ws.cell(row=current_row, column=1+i, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        current_row += 1
        
        # Sort statuses alphabetically and add data rows
        total_articles = len(status_list)
        for status in sorted(status_counter.keys()):
            count = status_counter[status]
            percentage = (count / total_articles * 100) if total_articles > 0 else 0
            
            # Status name
            status_cell = ws.cell(row=current_row, column=1, value=status)
            status_cell.border = border
            status_cell.alignment = left_alignment
            
            # Count
            count_cell = ws.cell(row=current_row, column=2, value=count)
            count_cell.border = border
            count_cell.alignment = center_alignment
            count_cell.fill = count_fill
            
            # Percentage
            pct_cell = ws.cell(row=current_row, column=3, value=f"{percentage:.1f}%")
            pct_cell.border = border
            pct_cell.alignment = center_alignment
            
            current_row += 1
        
        # Add total row
        total_row_cell = ws.cell(row=current_row, column=1, value="TOTAL")
        total_row_cell.font = header_font
        total_row_cell.border = border
        total_row_cell.fill = header_fill
        
        total_count_cell = ws.cell(row=current_row, column=2, value=total_articles)
        total_count_cell.font = header_font
        total_count_cell.border = border
        total_count_cell.fill = header_fill
        total_count_cell.alignment = center_alignment
        
        total_pct_cell = ws.cell(row=current_row, column=3, value="100.0%")
        total_pct_cell.font = header_font
        total_pct_cell.border = border
        total_pct_cell.fill = header_fill
        total_pct_cell.alignment = center_alignment
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 35
        
        # Save the workbook
        wb.save(excel_file_path)
        wb.close()
        
        print(f"\nSUCCESS: {sheet_name} sheet created successfully!")
        print(f"  Total articles: {total_articles}")
        print(f"  Unique statuses: {len(status_counter)}")
        for status in sorted(status_counter.keys()):
            print(f"    {status}: {status_counter[status]} articles")
        
        return True
        
    except Exception as e:
        print(f"ERROR creating status summary sheet: {e}")
        import traceback
        traceback.print_exc()
        return False


def diagnose_data_for_transition_analysis(excel_file_path, sample_count=5):
    """Diagnose the data structure to help debug transition analysis
    
    Args:
        excel_file_path (str): Path to the Excel file
        sample_count (int): Number of sample rows to analyze
    """
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(excel_file_path)
        ws = wb.active
        
        # Find column indices
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header:
                headers[header.lower().strip()] = col
        
        print("="*80)
        print("DATA DIAGNOSIS FOR TRANSITION ANALYSIS")
        print("="*80)
        print(f"Available columns: {list(headers.keys())}")
        
        required_columns = ['id', 'priority', 'status_reason', 'time spent']
        print(f"\\nRequired columns: {required_columns}")
        print(f"Missing columns: {[col for col in required_columns if col not in headers]}")
        
        print(f"\\nSample data (first {sample_count} rows):")
        print("-" * 100)
        
        for row in range(2, min(2 + sample_count, ws.max_row + 1)):
            article_id = ws.cell(row=row, column=headers.get('id', 1)).value
            priority = ws.cell(row=row, column=headers.get('priority', 1)).value
            status_reason = ws.cell(row=row, column=headers.get('status_reason', 1)).value
            time_spent = ws.cell(row=row, column=headers.get('time spent', 1)).value
            
            print(f"Row {row-1}:")
            print(f"  ID: {article_id}")
            print(f"  Priority: '{priority}' (type: {type(priority)})")
            print(f"  Status Reason: '{status_reason}' (type: {type(status_reason)})")
            print(f"  Time Spent: '{time_spent}' (type: {type(time_spent)})")
            print()
        
        # Analyze unique values and transitions
        priorities = set()
        status_reasons = set()
        article_transitions = defaultdict(list)
        
        for row in range(2, min(100, ws.max_row + 1)):  # Check more rows
            article_id = ws.cell(row=row, column=headers.get('id', 1)).value
            priority = ws.cell(row=row, column=headers.get('priority', 1)).value
            status_reason = ws.cell(row=row, column=headers.get('status_reason', 1)).value
            time_spent = ws.cell(row=row, column=headers.get('time spent', 1)).value
            
            if priority:
                priorities.add(str(priority).strip())
            if status_reason:
                status_reasons.add(str(status_reason).strip())
            
            # Track transitions per article
            if article_id and status_reason:
                # Use article_id as-is for reference
                article_transitions[str(article_id)].append({
                    'status': str(status_reason).strip(),
                    'time': str(time_spent).strip() if time_spent else '',
                    'priority': str(priority).strip() if priority else ''
                })
        
        print(f"Unique Priorities found: {sorted(priorities)}")
        print(f"Unique Status Reasons found: {sorted(status_reasons)}")
        
        # Show transition patterns for first few articles
        print("\\nTransition patterns (first 3 articles):")
        print("-" * 80)
        for i, (article_id, transitions) in enumerate(list(article_transitions.items())[:3]):
            print(f"Article {article_id}:")
            for j, trans in enumerate(transitions):
                status = trans['status']
                time = trans['time']
                priority = trans['priority']
                print(f"  {j+1}. Status: '{status}' | Time: '{time}' | Priority: '{priority}'")
        print("="*80)
        
        wb.close()
        
    except Exception as e:
        print(f"Error in diagnosis: {e}")


try:
    auth = HTTPKerberosAuth(mutual_authentication=OPTIONAL)
except Exception:
    auth = None
    print("⚠ Warning: Kerberos authentication not available")


def get_article_data(article_id):
    """Fetch article data from HSD-ES API"""
    url = f'{BASE_URL}/{article_id}'
    print(f"Fetching article: {article_id}...")
    
    try:
        response = requests.get(url, auth=auth)
        if response.status_code == 200:
            result = response.json()
            # API returns data in 'data' array
            if 'data' in result and len(result['data']) > 0:
                print(f"SUCCESS: Article {article_id} fetched successfully")
                return result['data'][0]
            else:
                print(f"ERROR: No data found in response")
                return None
        else:
            print(f"ERROR: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"ERROR Exception: {e}")
        return None


def get_status_transitions(article_id):
    """Fetch status transition history"""
    url = f'{BASE_URL}/{article_id}/statustransition'
    print(f"Fetching status transitions for: {article_id}...")
    
    try:
        response = requests.get(url, auth=auth)
        if response.status_code == 200:
            result = response.json()
            # Transitions are in result['status'] array
            if 'status' in result and isinstance(result['status'], list):
                transitions = result['status']
                print(f"SUCCESS: Found {len(transitions)} transitions")
                return transitions
            else:
                print(f"ERROR: No status transitions found")
                return []
        else:
            print(f"ERROR: {response.status_code} - {response.text}")
            return []
    except Exception as e:
        print(f"ERROR Exception: {e}")
        return []


def safe_get(data, key, default=''):
    """Safely get value from dictionary, handling nested keys"""
    if not data:
        return default
    
    if '.' in key:
        keys = key.split('.')
        value = data
        for k in keys:
            if isinstance(value, dict):
                value = value.get(k, {})
            else:
                return default
        return value if value != {} else default
    
    return data.get(key, default)


def create_excel(article_ids, output_filename='hsdes_export.xlsx'):
    """Create Excel file with article data and status transitions"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "HSD-ES Data"
    
    # Store article data for transition analysis (like the working script)
    article_data_list = []
    
    headers = [
        'sl_no', 'id', 'title', 'description', 'domain',
        'updated_date', 'updated_by', 'priority', 'reason', 'status',
        'status_reason', 'time_spent', 'family', 'component_affected', 'domain_affected'
    ]
    
    display_headers = [
        'sl. no.', 'id', 'title', 'description', 'domain',
        'updated_date', 'updated_by', 'priority', 'reason', 'status',
        'status_reason', 'time spent', 'family', 'component_affected', 'domain_affected'
    ]
    
    for col_num, header in enumerate(display_headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    current_row = 2
    
    for sl_no, article_id in enumerate(article_ids, 1):
        print(f"\n{'='*60}")
        print(f"Processing Article {sl_no}/{len(article_ids)}: {article_id}")
        print(f"{'='*60}")
        
        article_data = get_article_data(article_id)
        if not article_data:
            print(f"Skipping article {article_id} - no data")
            continue
        
        transitions = get_status_transitions(article_id)
        if not transitions:
            print(f"Warning: No transitions found for {article_id}")
            transitions = [{}]  # Create at least one row
        else:
            # Show latest transition first
            transitions = list(reversed(transitions))
        
        # Get current status from the latest transition (first in reversed list)
        current_status = 'Unknown'
        if transitions and isinstance(transitions[0], dict):
            current_status = transitions[0].get('status', 'Unknown')
        
        # Store article data for transition analysis (like the working script)
        article_analysis_data = {
            'id': article_data.get('id', ''),
            'priority': article_data.get('priority', ''),
            'status': current_status,  # Add current status
            'transitions': get_status_transitions(article_id)  # Get original order transitions
        }
        article_data_list.append(article_analysis_data)
        
        start_row = current_row
        
        article_fields = {
            'sl_no': sl_no,
            'id': article_data.get('id', ''),
            'title': article_data.get('title', ''),
            'description': article_data.get('description', ''),
            'domain': article_data.get('domain', ''),
            'priority': article_data.get('priority', ''),
            'family': article_data.get('family', ''),
            'component_affected': article_data.get('component_affected', ''),
            'domain_affected': article_data.get('domain_affected', '')
        }
        
        for trans_idx, transition in enumerate(transitions):
            row = current_row
            
            # Write article data only in first row (except priority)
            if trans_idx == 0:
                for col_num, header in enumerate(headers, 1):
                    if header in article_fields and header != 'priority':
                        cell = ws.cell(row=row, column=col_num, value=article_fields[header])
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                        
                        # Make the 'id' column a hyperlink
                        if header == 'id' and article_fields[header]:
                            article_id_value = article_fields[header]
                            cell.value = f'=HYPERLINK("https://hsdes.intel.com/resource/{article_id_value}", "{article_id_value}")'
            
            # Write priority in every row
            priority_col = headers.index('priority') + 1
            cell = ws.cell(row=row, column=priority_col, value=article_fields['priority'])
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if isinstance(transition, dict):
                # Extract status and reason from combined status (e.g., "open.debug" -> status="open", reason="debug")
                full_status = transition.get('status', '')
                if '.' in full_status:
                    status_parts = full_status.split('.', 1)
                    status = status_parts[0]
                    reason = status_parts[1]
                else:
                    status = full_status
                    reason = ''
                
                transition_fields = {
                    'updated_date': transition.get('updated_date', ''),
                    'updated_by': transition.get('updated_by', ''),
                    'reason': reason,
                    'status': status,
                    'status_reason': full_status,
                    'time_spent': transition.get('duration', '')
                }
                
                # If time_spent is empty and this is the first transition (most recent), 
                # calculate duration from transition date to current date
                if not transition_fields['time_spent'] and trans_idx == 0:
                    transition_fields['time_spent'] = calculate_duration_from_date(transition_fields['updated_date'])
            else:
                transition_fields = {
                    'updated_date': str(transition) if transition else '',
                    'updated_by': '',
                    'reason': '',
                    'status': '',
                    'status_reason': '',
                    'time_spent': ''
                }
            
            for col_num, header in enumerate(headers, 1):
                if header in transition_fields:
                    cell = ws.cell(row=row, column=col_num, value=transition_fields[header])
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
        
        if len(transitions) > 1:
            end_row = current_row - 1
            for col_num, header in enumerate(headers, 1):
                # Do NOT merge the 'priority' column - it should appear in every row
                if header in article_fields and header != 'priority':
                    col_letter = get_column_letter(col_num)
                    ws.merge_cells(f'{col_letter}{start_row}:{col_letter}{end_row}')
    
    for col_num in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 15
    
    ws.column_dimensions['A'].width = 8   # sl_no
    ws.column_dimensions['B'].width = 12  # id
    ws.column_dimensions['C'].width = 50  # title
    ws.column_dimensions['D'].width = 60  # description
    
    wb.save(output_filename)
    print(f"\n{'='*60}")
    print(f"SUCCESS: Excel file created: {output_filename}")
    print(f"{'='*60}")
    
    # Return article_data_list for transition analysis
    return article_data_list


def read_article_ids_from_excel(filename):
    """Read article IDs from Excel file (from 'id' column)"""
    try:
        from openpyxl import load_workbook
        import re
        
        wb = load_workbook(filename, data_only=False)  # Read formulas
        ws = wb.active
        
        article_ids = []
        
        id_col_index = None
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            if header and str(header).lower().strip() == 'id':
                id_col_index = col
                break
        
        if id_col_index is None:
            print("ERROR: 'id' column not found in Excel file!")
            wb.close()
            return None
        
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=id_col_index)
            cell_value = cell.value
            
            if cell_value:
                article_id = None
                article_id_str = str(cell_value).strip()
                
                # Handle HYPERLINK formulas - extract ID from the link text (second parameter)
                # Format: =HYPERLINK("url", "16027730784") or ==HYPERLINK("url", "16027730784")
                if 'HYPERLINK' in article_id_str:
                    # Extract the display text (second parameter in quotes)
                    match = re.search(r',\s*"([^"]+)"\s*\)', article_id_str)
                    if match:
                        article_id = match.group(1).strip()
                    else:
                        # Try to extract any number from the string
                        numbers = re.findall(r'\d+', article_id_str)
                        article_id = numbers[-1] if numbers else None
                else:
                    article_id = article_id_str
                
                if article_id:
                    # Keep only digits
                    article_id = re.sub(r'\D', '', article_id)
                    if article_id:
                        article_ids.append(article_id)
        
        wb.close()
        print(f"SUCCESS: Loaded {len(article_ids)} article IDs from '{filename}' (column: 'id')")
        return article_ids
        
    except FileNotFoundError:
        print(f"ERROR: File '{filename}' not found!")
        return None
    except Exception as e:
        print(f"ERROR reading Excel file: {e}")
        return None


def main():
    """Main function"""
    print("="*60)
    print("HSD-ES API Data Fetcher with Excel Export")
    print("="*60)
    print()
    
    print("Enter the Excel file path containing article IDs (with 'id' column)")
    filename = input("Excel file path: ").strip()
    filename = filename.strip('"').strip("'")  # Remove quotes if pasted
    
    if not filename:
        print("ERROR: No filename provided!")
        return
    
    article_ids = read_article_ids_from_excel(filename)
    if not article_ids:
        print("ERROR: No article IDs found or error reading file!")
        return
    
    print(f"\nArticle IDs to fetch: {', '.join(article_ids)}")
    
    default_filename = f"hsdes_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_filename = input(f"Output filename [{default_filename}]: ").strip()
    if not output_filename:
        output_filename = default_filename
    
    if not output_filename.endswith('.xlsx'):
        output_filename += '.xlsx'
    
    article_data_list = create_excel(article_ids, output_filename)
    
    print(f"\nSUCCESS: Done! Open '{output_filename}' in Excel to view the data.")
    
    # Ask if user wants to create transition analysis
    print("\n" + "="*60)
    create_analysis = input("Do you want to create Bugs Transition Graph? (y/n): ").strip().lower()
    if create_analysis in ['y', 'yes']:
        # Ask for platform name
        print("\nAvailable platforms: WCL, PTL-H, NVL-S, NVL-Hx, ARL-S-Ref, ARL-Hx-Ref")
        platform_name = input("Enter platform name: ").strip()
        if not platform_name:
            platform_name = "WCL"  # Default to WCL if empty
            print(f"No platform entered, defaulting to {platform_name}")
        
        print(f"Creating transition analysis for {platform_name}...")
        
        success = create_wcl_bugs_transition_graph_sheet(output_filename, article_data_list, platform_name)
        if success:
            print(f"\nSUCCESS: Transition analysis completed! Check the '{platform_name}_Bugs_transition_graph' sheet in '{output_filename}'")
        else:
            print("\nERROR: Failed to create transition analysis")
        
        # Also create status summary sheet
        print("\nCreating status summary sheet...")
        success2 = create_status_summary_sheet(output_filename, article_data_list, platform_name)
        if success2:
            print(f"SUCCESS: Status summary completed! Check the '{platform_name}_Status_Summary' sheet in '{output_filename}'")
        else:
            print("ERROR: Failed to create status summary sheet")


if __name__ == "__main__":
    main()
